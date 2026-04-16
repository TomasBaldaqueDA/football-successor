# -*- coding: utf-8 -*-
"""
Importa o CSV exportado do Supabase para a sheet DATA do dashboard.
- Multiplica colunas _pct_merged por 100 (passam de 0-1 para 0-100)
- Preserva todo o layout/formatação existente
"""
import sys, io, csv
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_PATH  = r"C:/Users/Lenovo/Desktop/Football_Successor_2026/artifacts/Supabase Snippet Player Profile Merged Stats Selector.csv"
XLSX_PATH = r"C:/Users/Lenovo/Desktop/Football_Successor_2026/artifacts/Player_Comparison_Dashboard.xlsx"

# ─── helpers ──────────────────────────────────────────────────────────────────
NAVY  = "0D1B2A"
SLATE = "1B2A3B"
WHITE = "FFFFFF"
MGRAY = "CBD5E1"
ZEBRA = "EFF6FF"

def thin(color=MGRAY):
    s = Side(border_style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def BG(h):
    return PatternFill("solid", fgColor=h)

# ─── Identify pct columns (0-based index in CSV) ──────────────────────────────
# Columns 21-29 (0-based): goals_p90_pct_merged … aerial_won_p90_pct_merged
# These come back as 0..1 from percent_rank() → multiply by 100
PCT_COL_NAMES = {
    "goals_p90_pct_merged", "assists_p90_pct_merged", "xg_per_90_pct_merged",
    "key_passes_p90_pct_merged", "tackles_p90_pct_merged",
    "interceptions_p90_pct_merged", "pass_success_pct_pct_merged",
    "dribbles_won_p90_pct_merged", "aerial_won_p90_pct_merged",
}

# ─── Read CSV ─────────────────────────────────────────────────────────────────
print("Reading CSV...")
with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
    reader = csv.reader(f)
    headers = next(reader)
    raw_rows = list(reader)

print(f"  {len(raw_rows)} players  x  {len(headers)} columns")

# Build per-column metadata
pct_idx = {i for i, h in enumerate(headers) if h in PCT_COL_NAMES}
numeric_idx = set(range(1, len(headers)))  # all except name (col 0) are numeric

def cast(val, col_idx):
    if val == "" or val is None:
        return None
    if col_idx in numeric_idx:
        try:
            return float(val)
        except ValueError:
            return val
    return val

# Apply pct * 100 during cast
def cast_pct(val, col_idx):
    v = cast(val, col_idx)
    if col_idx in pct_idx and isinstance(v, float):
        return round(v * 100, 2)
    return v

rows = [[cast_pct(cell, ci) for ci, cell in enumerate(row)] for row in raw_rows]

# ─── Load workbook ────────────────────────────────────────────────────────────
print("Loading workbook...")
wb = load_workbook(XLSX_PATH)
ws = wb["DATA"]

# ─── Find the header row (search for 'name' in col A) ────────────────────────
hdr_row = None
for r in ws.iter_rows(min_col=1, max_col=1):
    cell = r[0]
    if cell.value == "name":
        hdr_row = cell.row
        break

if hdr_row is None:
    raise RuntimeError("Could not find header row with 'name' in column A of DATA sheet.")

print(f"  Header row: {hdr_row}")

# Validate CSV columns match sheet headers
sheet_headers = [ws.cell(hdr_row, c + 1).value for c in range(len(headers))]
mismatches = [(i+1, headers[i], sheet_headers[i])
              for i in range(len(headers)) if headers[i] != sheet_headers[i]]
if mismatches:
    print("  COLUMN MISMATCHES:")
    for pos, csv_h, sheet_h in mismatches:
        print(f"    col {pos}: CSV='{csv_h}'  sheet='{sheet_h}'")
else:
    print("  All columns match perfectly")

# ─── Clear existing data rows (everything below hdr_row + 1) ─────────────────
DATA_START = hdr_row + 2   # +1 = divider row, +2 = first real data
print(f"  Clearing from row {DATA_START} …")

# Delete rows starting from DATA_START (work backwards to avoid index shift)
max_row = ws.max_row
if max_row >= DATA_START:
    ws.delete_rows(DATA_START, max_row - DATA_START + 1)

# ─── Write divider label ──────────────────────────────────────────────────────
div_row = hdr_row + 1
# (already written during build; just ensure it's there)

# ─── Write data rows ──────────────────────────────────────────────────────────
print(f"  Writing {len(rows)} rows starting at row {DATA_START}…")

NUM_FMT = {
    0: "@",          # name → text
    1: "0",          # age  → integer
    5: "0.00",       # rating
}
# raw p90 cols (6-20, 0-based = indices 6..20): 2 decimal places
for i in range(6, 21):
    NUM_FMT[i] = "0.000"
# pct cols (21-29, 0-based): 1 decimal (now 0-100)
for i in range(21, 30):
    NUM_FMT[i] = "0.0"

for ri, row in enumerate(rows):
    rn = DATA_START + ri
    bg = ZEBRA if ri % 2 == 0 else WHITE
    for ci, val in enumerate(row):
        c = ws.cell(rn, ci + 1, val)
        c.fill = BG(bg)
        c.border = thin()
        c.alignment = Alignment(
            horizontal="left" if ci == 0 else "center",
            vertical="center"
        )
        c.font = Font(name="Arial", size=9)
        fmt = NUM_FMT.get(ci)
        if fmt and val is not None:
            c.number_format = fmt

    if ri % 500 == 0 and ri > 0:
        print(f"    … {ri} rows written")

# ─── Update dropdown range to match actual data ───────────────────────────────
ws_dash = wb["DASHBOARD"]
data_end = DATA_START + len(rows) - 1
new_formula = f"DATA!$A${DATA_START}:$A${data_end}"

for dv in ws_dash.data_validations.dataValidation:
    if "DATA!$A$" in str(dv.formula1):
        dv.formula1 = new_formula
        print(f"  Updated dropdown: {new_formula}")

# ─── Save ─────────────────────────────────────────────────────────────────────
print("Saving workbook…")
wb.save(XLSX_PATH)
print(f"Done. {len(rows)} players loaded into {XLSX_PATH}")

# ─── Quick sanity check ───────────────────────────────────────────────────────
print("\nSample spot-check (first 3 players, pct cols):")
wb2 = load_workbook(XLSX_PATH, read_only=True)
ws2 = wb2["DATA"]
pct_col_letters = [get_column_letter(i + 1) for i in range(21, 30)]
for row in list(ws2.iter_rows(
        min_row=DATA_START, max_row=DATA_START + 2,
        min_col=1, max_col=30, values_only=True)):
    name = row[0]
    pct_vals = [round(row[i], 1) if row[i] is not None else None for i in range(21, 30)]
    print(f"  {name}: pct = {pct_vals}")
