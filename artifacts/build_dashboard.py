# -*- coding: utf-8 -*-
"""
Football Scout - Player Comparison Dashboard Builder
Source: mart.player_profile_merged_v1
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import RadarChart, BarChart, Reference, Series as ChSeries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

OUTPUT = r"C:/Users/Lenovo/Desktop/Football_Successor_2026/artifacts/Player_Comparison_Dashboard.xlsx"

# ─── Palette ──────────────────────────────────────────────────────────────────
NAVY   = "0D1B2A"
SLATE  = "1B2A3B"
P1     = "1A56DB"   # Blue   – Player 1
P2     = "C2410C"   # Orange – Player 2
LGRAY  = "F0F4F8"
MGRAY  = "CBD5E1"
DGRAY  = "64748B"
WHITE  = "FFFFFF"
ZEBRA  = "EFF6FF"
P1BG   = "EFF6FF"
P2BG   = "FFF7ED"

# ─── Helpers ──────────────────────────────────────────────────────────────────
def F(bold=False, size=10, color=NAVY, name="Arial", italic=False):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def BG(h):
    return PatternFill("solid", fgColor=h)

def AL(h="center", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def thin(color=MGRAY):
    s = Side(border_style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def bottom_only(color=MGRAY, thick=False):
    return Border(bottom=Side(border_style="medium" if thick else "thin", color=color))

def set_cell(ws, r, c, val, bold=False, size=10, color=NAVY, bg=None,
             align="center", wrap=False, italic=False, border=None, fmt=None):
    cell = ws.cell(r, c, val)
    cell.font = F(bold=bold, size=size, color=color, italic=italic)
    if bg:
        cell.fill = BG(bg)
    cell.alignment = AL(h=align, wrap=wrap)
    if border:
        cell.border = border
    if fmt:
        cell.number_format = fmt
    return cell

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTED COLUMNS  (order matches the SELECT in the SQL query below)
# ═══════════════════════════════════════════════════════════════════════════════
#
#  A  name
#  B  age
#  C  club
#  D  positions
#  E  nationality
#  F  rating_merged
#  G  goals_p90_merged
#  H  assists_p90_merged
#  I  xg_per_90_merged
#  J  shots_p90_merged
#  K  key_passes_p90_merged
#  L  passes_p90_merged
#  M  accurate_crosses_p90_merged
#  N  interceptions_p90_merged
#  O  tackles_p90_merged
#  P  blocks_p90_merged
#  Q  dribbles_won_p90_merged
#  R  aerial_won_p90_merged
#  S  pass_success_pct_merged
#  T  dispossessed_p90_merged
#  U  turnovers_p90_merged
#  V  goals_p90_pct_merged          ← pct 0-100
#  W  assists_p90_pct_merged
#  X  xg_per_90_pct_merged
#  Y  key_passes_p90_pct_merged
#  Z  tackles_p90_pct_merged
# AA  interceptions_p90_pct_merged
# AB  pass_success_pct_pct_merged
# AC  dribbles_won_p90_pct_merged
# AD  aerial_won_p90_pct_merged

DATA_COLS = [
    # (column_name_in_csv, short_label, col_width)
    ("name",                        "Name",         22),
    ("age",                         "Age",           6),
    ("club",                        "Club",         18),
    ("positions",                   "Pos",           8),
    ("nationality",                 "Nat",           7),
    ("rating_merged",               "Rating",        9),
    ("goals_p90_merged",            "G/90",          8),
    ("assists_p90_merged",          "A/90",          8),
    ("xg_per_90_merged",            "xG/90",         8),
    ("shots_p90_merged",            "Sh/90",         8),
    ("key_passes_p90_merged",       "KP/90",         8),
    ("passes_p90_merged",           "P/90",          8),
    ("accurate_crosses_p90_merged", "Crs/90",        8),
    ("interceptions_p90_merged",    "Int/90",        8),
    ("tackles_p90_merged",          "Tkl/90",        8),
    ("blocks_p90_merged",           "Blk/90",        8),
    ("dribbles_won_p90_merged",     "Drb/90",        8),
    ("aerial_won_p90_merged",       "Aer/90",        8),
    ("pass_success_pct_merged",     "Pass%",         8),
    ("dispossessed_p90_merged",     "Disp/90",       9),
    ("turnovers_p90_merged",        "TO/90",         8),
    ("goals_p90_pct_merged",        "G% pct",        8),
    ("assists_p90_pct_merged",      "A% pct",        8),
    ("xg_per_90_pct_merged",        "xG% pct",       9),
    ("key_passes_p90_pct_merged",   "KP% pct",       9),
    ("tackles_p90_pct_merged",      "Tkl% pct",      9),
    ("interceptions_p90_pct_merged","Int% pct",      9),
    ("pass_success_pct_pct_merged", "Pass%pct",      9),
    ("dribbles_won_p90_pct_merged", "Drb% pct",      9),
    ("aerial_won_p90_pct_merged",   "Aer% pct",      9),
]

SQL_QUERY = [
    "SELECT",
    "  name, age, club, positions, nationality, rating_merged,",
    "  goals_p90_merged, assists_p90_merged, xg_per_90_merged, shots_p90_merged,",
    "  key_passes_p90_merged, passes_p90_merged, accurate_crosses_p90_merged,",
    "  interceptions_p90_merged, tackles_p90_merged, blocks_p90_merged,",
    "  dribbles_won_p90_merged, aerial_won_p90_merged, pass_success_pct_merged,",
    "  dispossessed_p90_merged, turnovers_p90_merged,",
    "  goals_p90_pct_merged, assists_p90_pct_merged, xg_per_90_pct_merged,",
    "  key_passes_p90_pct_merged, tackles_p90_pct_merged, interceptions_p90_pct_merged,",
    "  pass_success_pct_pct_merged, dribbles_won_p90_pct_merged, aerial_won_p90_pct_merged",
    "FROM mart.player_profile_merged_v1",
    "WHERE name IS NOT NULL",
    "ORDER BY name ASC;",
]

# Percentile columns used for RADAR (0-100 scale, already percentile-ranked)
RADAR_METRICS = [
    ("Goals",      "V"),   # goals_p90_pct_merged
    ("xG",         "X"),   # xg_per_90_pct_merged
    ("Key Passes", "Y"),   # key_passes_p90_pct_merged
    ("Tackles",    "Z"),   # tackles_p90_pct_merged
    ("Pass %",     "AB"),  # pass_success_pct_pct_merged
    ("Dribbles",   "AC"),  # dribbles_won_p90_pct_merged
]

# Raw p90 columns used for BAR chart
BAR_METRICS = [
    ("Goals/90",    "G"),
    ("Assists/90",  "H"),
    ("xG/90",       "I"),
    ("Shots/90",    "J"),
    ("Key Pass/90", "K"),
    ("Passes/90",   "L"),
    ("Crosses/90",  "M"),
    ("Interc./90",  "N"),
    ("Tackles/90",  "O"),
    ("Blocks/90",   "P"),
    ("Dribbles/90", "Q"),
    ("Aerials/90",  "R"),
]

# Percentile KPIs shown as score cards (uses same pct cols as RADAR)
SCORECARD_METRICS = [
    ("Scoring",    "V"),   # goals pct
    ("Creating",   "Y"),   # key passes pct
    ("Defending",  "Z"),   # tackles pct
    ("Possession", "AB"),  # pass success pct
    ("Dribbling",  "AC"),  # dribbles pct
]

# Sample rows (name, age, club, positions, nationality, rating, + raw cols G-U, + pct cols V-AD)
SAMPLES = [
    ["Bruno Fernandes", 29, "Man United", "CAM", "PRT", 7.8,
     0.18, 0.42, 0.21, 2.8, 3.10, 45.2, 1.2, 0.8, 1.2, 0.3, 0.9, 0.4, 86.3, 0.9, 0.7,
     64.0, 72.0, 70.0, 83.0, 38.0, 42.0, 75.0, 58.0, 44.0, 51.0],
    ["Bernardo Silva", 29, "Man City", "CM", "PRT", 8.1,
     0.11, 0.31, 0.14, 1.6, 4.20, 54.8, 1.4, 1.3, 1.4, 0.5, 1.2, 0.3, 89.1, 0.6, 0.5,
     48.0, 52.0, 50.0, 90.0, 56.0, 60.0, 88.0, 74.0, 52.0, 38.0],
    ["Martin Odegaard", 26, "Arsenal", "CAM", "NOR", 8.3,
     0.19, 0.35, 0.22, 2.1, 3.90, 48.3, 1.1, 0.9, 1.1, 0.4, 0.8, 0.2, 88.4, 0.7, 0.6,
     70.0, 72.0, 74.0, 92.0, 40.0, 45.0, 84.0, 65.0, 48.0, 27.0],
]

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: DATA
# ═══════════════════════════════════════════════════════════════════════════════
ws_data = wb.active
ws_data.title = "DATA"
ws_data.sheet_properties.tabColor = DGRAY

# Instructions banner (rows 1-2)
ws_data.merge_cells("A1:AD1")
inst = ws_data.cell(1, 1,
    "INSTRUCOES  |  Exporta o CSV do Supabase com a query abaixo. "
    "No Excel: Dados -> Obter Dados -> Do Ficheiro -> Do CSV -> "
    "carregar nesta sheet substituindo os dados existentes (linha 10 em diante).")
inst.font = F(size=9, color=WHITE)
inst.fill = BG(NAVY)
inst.alignment = AL(h="left", v="center", wrap=True, indent=1)
ws_data.row_dimensions[1].height = 36

for i, line in enumerate(SQL_QUERY, 2):
    ws_data.merge_cells(f"A{i}:AD{i}")
    c = ws_data.cell(i, 1, line)
    c.font = Font(name="Courier New", size=8, color="1E3A5F")
    c.fill = BG("F0F7FF")
    c.alignment = AL(h="left", indent=1)
    ws_data.row_dimensions[i].height = 13

# Header row — must be BELOW the SQL block (SQL occupies rows 2 … 2+len-1)
HDR_ROW = 2 + len(SQL_QUERY) + 2   # +1 blank spacer row
for ci, (col_name, label, width) in enumerate(DATA_COLS, 1):
    c = ws_data.cell(HDR_ROW, ci, col_name)
    c.font = F(bold=True, size=8, color=WHITE)
    c.fill = BG(SLATE)
    c.alignment = AL()
    c.border = thin()
    ws_data.column_dimensions[get_column_letter(ci)].width = width

# Divider label row below header
lbl_row = HDR_ROW + 1
ws_data.merge_cells(f"A{lbl_row}:F{lbl_row}")
ws_data.cell(lbl_row, 1, "-- DADOS DO SUPABASE ABAIXO --").font = F(size=7, color=DGRAY, italic=True)
ws_data.cell(lbl_row, 1).fill = BG("F8FAFC")
ws_data.cell(lbl_row, 1).alignment = AL(h="center")

# Sample rows
DATA_START = HDR_ROW + 2
for ri, row in enumerate(SAMPLES):
    rn = DATA_START + ri
    bg = ZEBRA if ri % 2 == 0 else WHITE
    for ci, val in enumerate(row, 1):
        c = ws_data.cell(rn, ci, val)
        c.font = F(size=9)
        c.fill = BG(bg)
        c.alignment = AL(h="left" if ci == 1 else "center")
        c.border = thin()
        if ci == 6:
            c.number_format = "0.0"
        elif ci in range(7, 22):
            c.number_format = "0.00"
        elif ci in range(22, 31):
            c.number_format = "0.0"

ws_data.freeze_panes = "B11"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: RADAR_DATA
# ═══════════════════════════════════════════════════════════════════════════════
ws_radar = wb.create_sheet("RADAR_DATA")
ws_radar.sheet_properties.tabColor = "334155"

set_cell(ws_radar, 1, 1, "Metric",   bold=True, bg=NAVY,  color=WHITE, size=9)
set_cell(ws_radar, 1, 2, "Player 1", bold=True, bg=P1, color=WHITE, size=9)
set_cell(ws_radar, 1, 3, "Player 2", bold=True, bg=P2, color=WHITE, size=9)
ws_radar.column_dimensions["A"].width = 14
ws_radar.column_dimensions["B"].width = 12
ws_radar.column_dimensions["C"].width = 12

# name lookup cells: DASHBOARD!$B$6 = P1 name, DASHBOARD!$I$6 = P2 name
for i, (label, dcol) in enumerate(RADAR_METRICS, 2):
    set_cell(ws_radar, i, 1, label, bold=True, align="left",
             bg=LGRAY, size=9)
    ws_radar.cell(i, 2).value = (
        f"=IFERROR(XLOOKUP(DASHBOARD!$B$6,"
        f"DATA!$A:$A,DATA!${dcol}:${dcol},0),0)"
    )
    ws_radar.cell(i, 2).number_format = "0.0"
    ws_radar.cell(i, 2).font = F(size=9)
    ws_radar.cell(i, 2).alignment = AL()
    ws_radar.cell(i, 3).value = (
        f"=IFERROR(XLOOKUP(DASHBOARD!$I$6,"
        f"DATA!$A:$A,DATA!${dcol}:${dcol},0),0)"
    )
    ws_radar.cell(i, 3).number_format = "0.0"
    ws_radar.cell(i, 3).font = F(size=9)
    ws_radar.cell(i, 3).alignment = AL()

# Radar chart
radar = RadarChart()
radar.type = "filled"
radar.style = 26
radar.title = None
radar.legend = None
radar.shape = 0

cats    = Reference(ws_radar, min_col=1, min_row=2, max_row=7)
s1_ref  = Reference(ws_radar, min_col=2, min_row=1, max_row=7)
s2_ref  = Reference(ws_radar, min_col=3, min_row=1, max_row=7)

s1 = ChSeries(s1_ref, title_from_data=True)
s1.graphicalProperties.line.solidFill = P1
s1.graphicalProperties.line.width = 22000
s1.graphicalProperties.solidFill = P1 + "44"

s2 = ChSeries(s2_ref, title_from_data=True)
s2.graphicalProperties.line.solidFill = P2
s2.graphicalProperties.line.width = 22000
s2.graphicalProperties.solidFill = P2 + "44"

radar.append(s1)
radar.append(s2)
radar.set_categories(cats)
radar.width  = 14
radar.height = 14
ws_radar.add_chart(radar, "E2")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: BAR_DATA
# ═══════════════════════════════════════════════════════════════════════════════
ws_bar = wb.create_sheet("BAR_DATA")
ws_bar.sheet_properties.tabColor = "334155"

set_cell(ws_bar, 1, 1, "Metric",   bold=True, bg=NAVY, color=WHITE, size=9)
set_cell(ws_bar, 1, 2, "Player 1", bold=True, bg=P1,   color=WHITE, size=9)
set_cell(ws_bar, 1, 3, "Player 2", bold=True, bg=P2,   color=WHITE, size=9)
ws_bar.column_dimensions["A"].width = 14
ws_bar.column_dimensions["B"].width = 11
ws_bar.column_dimensions["C"].width = 11

for i, (label, dcol) in enumerate(BAR_METRICS, 2):
    set_cell(ws_bar, i, 1, label, bold=True, align="left", bg=LGRAY, size=9)
    ws_bar.cell(i, 2).value = (
        f"=IFERROR(XLOOKUP(DASHBOARD!$B$6,"
        f"DATA!$A:$A,DATA!${dcol}:${dcol},0),0)"
    )
    ws_bar.cell(i, 2).number_format = "0.00"
    ws_bar.cell(i, 2).font = F(size=9)
    ws_bar.cell(i, 2).alignment = AL()
    ws_bar.cell(i, 3).value = (
        f"=IFERROR(XLOOKUP(DASHBOARD!$I$6,"
        f"DATA!$A:$A,DATA!${dcol}:${dcol},0),0)"
    )
    ws_bar.cell(i, 3).number_format = "0.00"
    ws_bar.cell(i, 3).font = F(size=9)
    ws_bar.cell(i, 3).alignment = AL()

N_BAR = len(BAR_METRICS)

bar = BarChart()
bar.type = "bar"
bar.barDir = "bar"
bar.grouping = "clustered"
bar.title = None
bar.legend = None
bar.style = 10
bar.overlap = -10

bar_cats = Reference(ws_bar, min_col=1, min_row=2, max_row=1 + N_BAR)
b1_ref   = Reference(ws_bar, min_col=2, min_row=1, max_row=1 + N_BAR)
b2_ref   = Reference(ws_bar, min_col=3, min_row=1, max_row=1 + N_BAR)

b1 = ChSeries(b1_ref, title_from_data=True)
b1.graphicalProperties.solidFill = P1
b2 = ChSeries(b2_ref, title_from_data=True)
b2.graphicalProperties.solidFill = P2

bar.append(b1)
bar.append(b2)
bar.set_categories(bar_cats)
bar.width  = 17
bar.height = 18
ws_bar.add_chart(bar, "E2")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DASHBOARD", 0)
ws.sheet_properties.tabColor = P1
ws.sheet_view.showGridLines = False

# Column widths
COL_W = {
    "A": 1.5, "B": 2,   "C": 18,  "D": 14,  "E": 2,
    "F": 2,   "G": 2,   "H": 18,  "I": 14,  "J": 2,
    "K": 2,   "L": 14,  "M": 14,  "N": 2,
}
for col, w in COL_W.items():
    ws.column_dimensions[col].width = w

# Row heights
ROW_H = {
    1: 5,   2: 38,  3: 6,   4: 26,  5: 6,
    6: 22,  7: 22,  8: 22,  9: 22,  10: 22,
    11: 6,  12: 26, 13: 6,
    14: 20, 15: 28, 16: 28, 17: 22, 18: 6,
}
for r, h in ROW_H.items():
    ws.row_dimensions[r].height = h

# ── TITLE ─────────────────────────────────────────────────────────────────────
ws.merge_cells("B2:N2")
tc = ws.cell(2, 2, "  FOOTBALL SCOUTING   |   PLAYER COMPARISON DASHBOARD")
tc.font = Font(name="Arial", bold=True, size=17, color=WHITE)
tc.fill = BG(NAVY)
tc.alignment = AL(h="left", v="center")

# ── PLAYER HEADERS (row 4) ────────────────────────────────────────────────────
ws.merge_cells("B4:D4")
ws.cell(4, 2, "PLAYER  1").font  = F(bold=True, size=10, color=WHITE)
ws.cell(4, 2).fill       = BG(P1)
ws.cell(4, 2).alignment  = AL()

ws.merge_cells("F4:G4")
vs = ws.cell(4, 6, "VS")
vs.font = Font(name="Arial", bold=True, size=14, color=NAVY)
vs.alignment = AL()

ws.merge_cells("H4:J4")
ws.cell(4, 8, "PLAYER  2").font  = F(bold=True, size=10, color=WHITE)
ws.cell(4, 8).fill       = BG(P2)
ws.cell(4, 8).alignment  = AL()

# ── DROPDOWNS (row 5 → B5 and I5, merged) ────────────────────────────────────
# Note: player name lookup uses B5 and I5 as anchor
ws.merge_cells("B5:D5")  # Player 1 name cell  → col B = col index 2
ws.cell(5, 2, SAMPLES[0][0]).font = F(bold=True, size=12, color=P1)
ws.cell(5, 2).fill = BG(P1BG)
ws.cell(5, 2).alignment = AL()
ws.cell(5, 2).border = Border(bottom=Side(border_style="medium", color=P1))

ws.merge_cells("H5:J5")  # Player 2 name cell  → col H = col index 8
ws.cell(5, 8, SAMPLES[1][0]).font = F(bold=True, size=12, color=P2)
ws.cell(5, 8).fill = BG(P2BG)
ws.cell(5, 8).alignment = AL()
ws.cell(5, 8).border = Border(bottom=Side(border_style="medium", color=P2))

dv1 = DataValidation(type="list", formula1=f"DATA!$A${HDR_ROW+2}:$A$5000",
                     showDropDown=False, showErrorMessage=False, showInputMessage=False)
ws.add_data_validation(dv1)
dv1.add("B5")

dv2 = DataValidation(type="list", formula1=f"DATA!$A${HDR_ROW+2}:$A$5000",
                     showDropDown=False, showErrorMessage=False, showInputMessage=False)
ws.add_data_validation(dv2)
dv2.add("H5")

# ── INFO ROWS (6-10) ──────────────────────────────────────────────────────────
# Cols in DATA: club=C(3), positions=D(4), nationality=E(5), age=B(2), rating=F(6)
INFO_FIELDS = [
    ("Club",     "DATA!$C:$C"),
    ("Position", "DATA!$D:$D"),
    ("Nat.",     "DATA!$E:$E"),
    ("Age",      "DATA!$B:$B"),
    ("Rating",   "DATA!$F:$F"),
]

for idx, (label, dcol) in enumerate(INFO_FIELDS):
    row = 6 + idx
    bg = ZEBRA if idx % 2 == 0 else WHITE

    # label
    lbl = ws.cell(row, 2, label)
    lbl.font = F(size=8, color=DGRAY, bold=True)
    lbl.alignment = AL(h="right")

    # P1 value (col C = 3)
    ws.merge_cells(f"C{row}:D{row}")
    p1v = ws.cell(row, 3)
    if label == "Rating":
        p1v.value = f"=IFERROR(ROUND(XLOOKUP($B$5,DATA!$A:$A,{dcol},0),1),\"--\")"
        p1v.number_format = "0.0"
    else:
        p1v.value = f"=IFERROR(XLOOKUP($B$5,DATA!$A:$A,{dcol},\"--\"),\"--\")"
    p1v.font = F(size=9, bold=(label == "Rating"))
    p1v.fill = BG(bg)
    p1v.alignment = AL()

    # P2 value (col I = 9)
    ws.merge_cells(f"I{row}:J{row}")
    p2v = ws.cell(row, 9)
    if label == "Rating":
        p2v.value = f"=IFERROR(ROUND(XLOOKUP($H$5,DATA!$A:$A,{dcol},0),1),\"--\")"
        p2v.number_format = "0.0"
    else:
        p2v.value = f"=IFERROR(XLOOKUP($H$5,DATA!$A:$A,{dcol},\"--\"),\"--\")"
    p2v.font = F(size=9, bold=(label == "Rating"))
    p2v.fill = BG(bg)
    p2v.alignment = AL()

# ── SCORECARD SECTION (rows 12-17) ───────────────────────────────────────────
ws.merge_cells("B12:N12")
sc_hdr = ws.cell(12, 2, "PERCENTILE SCORES  ( 0 = lowest  /  100 = highest in database )")
sc_hdr.font = F(bold=True, size=9, color=WHITE, italic=True)
sc_hdr.fill = BG(SLATE)
sc_hdr.alignment = AL(h="left", indent=2)

# Each scorecard occupies 2 columns: B-C, D-E, F-G, H-I, J-K
SC_COL_START = 2  # col B
SC_COL_SPAN  = 2

for si, (sc_label, dcol) in enumerate(SCORECARD_METRICS):
    col_start = SC_COL_START + si * SC_COL_SPAN   # B=2, D=4, F=6, H=8, J=10
    col_end   = col_start + SC_COL_SPAN - 1

    cs  = get_column_letter(col_start)
    ce  = get_column_letter(col_end)

    # metric label row 14
    ws.merge_cells(f"{cs}14:{ce}14")
    ml = ws.cell(14, col_start, sc_label)
    ml.font = F(bold=True, size=8, color=DGRAY)
    ml.alignment = AL()
    ml.fill = BG(LGRAY)

    # P1 value row 15
    ws.merge_cells(f"{cs}15:{ce}15")
    p1s = ws.cell(15, col_start)
    p1s.value = (
        f"=IFERROR(ROUND(XLOOKUP($B$5,DATA!$A:$A,DATA!${dcol}:${dcol},0),1),\"--\")"
    )
    p1s.font = F(bold=True, size=13, color=P1)
    p1s.fill = BG(P1BG)
    p1s.alignment = AL()
    p1s.number_format = "0.0"

    # P2 value row 16
    ws.merge_cells(f"{cs}16:{ce}16")
    p2s = ws.cell(16, col_start)
    p2s.value = (
        f"=IFERROR(ROUND(XLOOKUP($H$5,DATA!$A:$A,DATA!${dcol}:${dcol},0),1),\"--\")"
    )
    p2s.font = F(bold=True, size=13, color=P2)
    p2s.fill = BG(P2BG)
    p2s.alignment = AL()
    p2s.number_format = "0.0"

    # Delta row 17 (P1 - P2)
    ws.merge_cells(f"{cs}17:{ce}17")
    dlt = ws.cell(17, col_start)
    dlt.value = (
        f"=IFERROR(ROUND("
        f"XLOOKUP($B$5,DATA!$A:$A,DATA!${dcol}:${dcol},0)"
        f"-XLOOKUP($H$5,DATA!$A:$A,DATA!${dcol}:${dcol},0)"
        f",1),\"--\")"
    )
    dlt.font = F(bold=True, size=10, color=NAVY)
    dlt.alignment = AL()
    dlt.number_format = "+0.0;-0.0;0.0"

    # Conditional formatting on delta
    cell_addr = f"{cs}17"
    ws.conditional_formatting.add(
        cell_addr,
        CellIsRule(operator="greaterThan", formula=["0"],
                   font=Font(color="065F46", bold=True))
    )
    ws.conditional_formatting.add(
        cell_addr,
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(color="9B1C1C", bold=True))
    )

# Delta label
ws.cell(17, 12, "P1 - P2").font = F(size=7, color=DGRAY, italic=True)
ws.cell(17, 12).alignment = AL(h="left")

# ── RE-BUILD CHARTS ANCHORED IN DASHBOARD ────────────────────────────────────
# Radar
radar2 = RadarChart()
radar2.type = "filled"
radar2.style = 26
radar2.title = None
radar2.legend = None
radar2.shape = 0

cats2  = Reference(ws_radar, min_col=1, min_row=2, max_row=7)
s1r    = Reference(ws_radar, min_col=2, min_row=1, max_row=7)
s2r    = Reference(ws_radar, min_col=3, min_row=1, max_row=7)
s1n = ChSeries(s1r, title_from_data=True)
s2n = ChSeries(s2r, title_from_data=True)
s1n.graphicalProperties.line.solidFill = P1
s1n.graphicalProperties.line.width = 22000
s1n.graphicalProperties.solidFill = P1 + "44"
s2n.graphicalProperties.line.solidFill = P2
s2n.graphicalProperties.line.width = 22000
s2n.graphicalProperties.solidFill = P2 + "44"
radar2.append(s1n)
radar2.append(s2n)
radar2.set_categories(cats2)
radar2.width  = 12
radar2.height = 13
ws.add_chart(radar2, "B19")

# Bar
bar2 = BarChart()
bar2.type = "bar"
bar2.barDir = "bar"
bar2.grouping = "clustered"
bar2.title = None
bar2.legend = None
bar2.style = 10
bar2.overlap = -10

barcats2 = Reference(ws_bar, min_col=1, min_row=2, max_row=1 + N_BAR)
b1r = Reference(ws_bar, min_col=2, min_row=1, max_row=1 + N_BAR)
b2r = Reference(ws_bar, min_col=3, min_row=1, max_row=1 + N_BAR)
b1n = ChSeries(b1r, title_from_data=True)
b2n = ChSeries(b2r, title_from_data=True)
b1n.graphicalProperties.solidFill = P1
b2n.graphicalProperties.solidFill = P2
bar2.append(b1n)
bar2.append(b2n)
bar2.set_categories(barcats2)
bar2.width  = 15
bar2.height = 17
ws.add_chart(bar2, "H19")

# ── SECTION LABELS above charts ───────────────────────────────────────────────
ws.merge_cells("B18:G18")
rl = ws.cell(18, 2, "RADAR  |  Percentile scores (0 - 100)")
rl.font = F(bold=True, size=8, color=WHITE)
rl.fill = BG(SLATE)
rl.alignment = AL(h="left", indent=2)

ws.merge_cells("H18:N18")
bl = ws.cell(18, 8, "METRICS  |  Values per 90 min")
bl.font = F(bold=True, size=8, color=WHITE)
bl.fill = BG(SLATE)
bl.alignment = AL(h="left", indent=2)

# ── LEGEND + FOOTER ──────────────────────────────────────────────────────────
FOOT = 50
ws.merge_cells(f"B{FOOT}:D{FOOT}")
l1 = ws.cell(FOOT, 2, "  Player 1")
l1.font = F(bold=True, size=9, color=P1)
l1.fill = BG(P1BG)
l1.alignment = AL(h="center")

ws.merge_cells(f"H{FOOT}:J{FOOT}")
l2 = ws.cell(FOOT, 8, "  Player 2")
l2.font = F(bold=True, size=9, color=P2)
l2.fill = BG(P2BG)
l2.alignment = AL(h="center")

ws.merge_cells(f"E{FOOT}:G{FOOT}")
note = ws.cell(FOOT, 5, "Radar = percentile vs all players in DB")
note.font = F(size=7, color=DGRAY, italic=True)
note.alignment = AL(h="center")

FOOT2 = FOOT + 1
ws.merge_cells(f"B{FOOT2}:N{FOOT2}")
foot = ws.cell(FOOT2, 2,
    "Football Successor 2026  |  Source: mart.player_profile_merged_v1  |  "
    "Atualizar: Dados -> Atualizar Tudo")
foot.font = F(size=7, color=MGRAY, italic=True)
foot.fill = BG(NAVY)
foot.alignment = AL(h="center")

# ── Freeze panes ──────────────────────────────────────────────────────────────
ws.freeze_panes = "B5"

# ── Hide auxiliary sheets ─────────────────────────────────────────────────────
ws_radar.sheet_state = "hidden"
ws_bar.sheet_state   = "hidden"

# ─────────────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
